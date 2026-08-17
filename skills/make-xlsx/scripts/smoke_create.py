#!/usr/bin/env python3
"""Smoke: create a sample .xlsx and run check_xlsx_basics.py.

Usage (from repo / workspace root):
  python skills/make-xlsx/scripts/smoke_create.py
"""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SCRIPT_DIR = Path(__file__).resolve().parent
WORKSPACE = SCRIPT_DIR.parents[2]
OUT = WORKSPACE / "outputs" / "documents" / "xlsx" / "smoke-product-mix.xlsx"


def main() -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(name="Calibri", bold=True, color="1A1A1A")
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    headers = ["Item", "Amount ($)", "Share"]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(1, col, text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    ws["A2"] = "Product A"
    ws["B2"] = 1200
    ws["A3"] = "Product B"
    ws["B3"] = 800
    ws["A4"] = "Total"
    ws["B4"] = "=SUM(B2:B3)"
    ws["C2"] = "=B2/$B$4"
    ws["C3"] = "=B3/$B$4"

    for coord in ("B2", "B3", "B4"):
        ws[coord].number_format = "$#,##0"
    for coord in ("C2", "C3"):
        ws[coord].number_format = "0.0%"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.freeze_panes = "A2"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}", flush=True)

    checker = SCRIPT_DIR / "check_xlsx_basics.py"
    proc = subprocess.run(
        [sys.executable, str(checker), str(OUT)],
        check=False,
    )
    return proc.returncode


if __name__ == "__main__":
    raise SystemExit(main())
