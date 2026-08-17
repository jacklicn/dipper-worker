#!/usr/bin/env python3
"""Maintainer smoke helper: XLSX package + basic structure (openpyxl).

Not part of the agent delivery path — agents must write valid workbooks via
openpyxl/pandas on first save. Do not use LibreOffice to “fix” files.
This script only supports local smoke_create.

Usage:
  python check_xlsx_basics.py book.xlsx
  python check_xlsx_basics.py book.xlsx --json

Exit 0 when status is ok (warnings allowed); exit 1 on errors.
"""

from __future__ import annotations

import argparse
import json
import re
import zipfile
from pathlib import Path
from urllib.parse import unquote
from xml.etree import ElementTree as ET

ERROR_VALUES = {
    "#REF!",
    "#DIV/0!",
    "#VALUE!",
    "#N/A",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#GETTING_DATA",
}

FORMULA_RE = re.compile(r"^=.+")

REQUIRED_PARTS = (
    "[Content_Types].xml",
    "xl/workbook.xml",
    "xl/_rels/workbook.xml.rels",
)


def local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def normalize_zip_name(name: str) -> str:
    return name.replace("\\", "/").lstrip("/")


def resolve_rel_target(base_dir: str, target: str) -> str:
    t = unquote(target.replace("\\", "/"))
    if t.startswith("/"):
        return t.lstrip("/")
    parts = [p for p in base_dir.split("/") if p] + [p for p in t.split("/") if p]
    out: list[str] = []
    for p in parts:
        if p == "..":
            if out:
                out.pop()
        elif p != ".":
            out.append(p)
    return "/".join(out)


def collect_internal_rel_targets(zf: zipfile.ZipFile, names: set[str]) -> set[str]:
    targets: set[str] = set()
    for name in names:
        if not name.endswith(".rels") or "/_rels/" not in name:
            continue
        parent = name.rpartition("/_rels/")[0]
        try:
            root = ET.fromstring(zf.read(name))
        except (KeyError, ET.ParseError):
            continue
        for el in root:
            if local(el.tag) != "Relationship":
                continue
            target = el.attrib.get("Target")
            if not target:
                continue
            mode = (el.attrib.get("TargetMode") or "Internal").lower()
            if mode == "external":
                continue
            targets.add(resolve_rel_target(parent, target))
    return targets


def check_package(path: Path) -> tuple[list[dict], list[dict]]:
    errors: list[dict] = []
    warnings: list[dict] = []
    try:
        zf = zipfile.ZipFile(path, "r")
    except zipfile.BadZipFile as e:
        return [{"message": f"not a valid ZIP/OOXML package: {e}"}], []

    with zf:
        names = {normalize_zip_name(n) for n in zf.namelist()}
        for part in REQUIRED_PARTS:
            if part not in names:
                errors.append({"message": f"missing required part: {part}"})

        referenced = collect_internal_rel_targets(zf, names)

        if "[Content_Types].xml" in names:
            try:
                ct_root = ET.fromstring(zf.read("[Content_Types].xml"))
                has_wb = False
                for el in ct_root:
                    if local(el.tag) != "Override":
                        continue
                    part_name = normalize_zip_name(el.attrib.get("PartName", "")).lstrip("/")
                    if part_name == "xl/workbook.xml":
                        has_wb = True
                    if part_name and part_name not in names:
                        entry = {
                            "message": "Content_Types Override points to missing part",
                            "part": part_name,
                        }
                        if part_name in referenced:
                            errors.append(entry)
                        else:
                            warnings.append(entry)
                if not has_wb and "xl/workbook.xml" in names:
                    errors.append(
                        {"message": "Content_Types missing Override for xl/workbook.xml"}
                    )
            except ET.ParseError as e:
                errors.append({"message": f"[Content_Types].xml parse error: {e}"})

        rels_path = "xl/_rels/workbook.xml.rels"
        if rels_path in names:
            try:
                root = ET.fromstring(zf.read(rels_path))
            except ET.ParseError as e:
                errors.append({"message": f"workbook.xml.rels parse error: {e}"})
            else:
                sheet_targets = 0
                for el in root:
                    if local(el.tag) != "Relationship":
                        continue
                    target = el.attrib.get("Target")
                    if not target:
                        errors.append({"message": "relationship missing Target"})
                        continue
                    mode = (el.attrib.get("TargetMode") or "Internal").lower()
                    if mode == "external":
                        continue
                    resolved = resolve_rel_target("xl", target)
                    rel_type = el.attrib.get("Type", "")
                    if "worksheet" in rel_type.lower() or "/sheet" in target.lower():
                        sheet_targets += 1
                    if resolved not in names:
                        errors.append(
                            {
                                "message": "broken relationship target",
                                "id": el.attrib.get("Id"),
                                "target": target,
                                "resolved": resolved,
                            }
                        )
                if sheet_targets == 0:
                    errors.append(
                        {"message": "workbook.xml.rels has no worksheet relationships"}
                    )

    return errors, warnings


def check_file(path: Path) -> dict:
    package_errors, package_warnings = check_package(path)

    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            "status": "errors_found",
            "file": str(path),
            "error_count": 1 + len(package_errors),
            "errors": [{"message": "openpyxl required: pip install openpyxl"}]
            + package_errors,
            "warnings": package_warnings,
        }

    errors: list[dict] = list(package_errors)
    warnings: list[dict] = list(package_warnings)
    sheets_info: list[dict] = []
    formula_count = 0
    error_cells = 0

    try:
        wb = load_workbook(path, data_only=False)
    except Exception as e:
        errors.append({"message": f"cannot open workbook: {e}"})
        return {
            "status": "errors_found",
            "file": str(path),
            "error_count": len(errors),
            "errors": errors,
            "warnings": warnings,
        }

    if not wb.sheetnames:
        errors.append({"message": "workbook has no sheets"})

    non_empty = 0
    for name in wb.sheetnames:
        ws = wb[name]
        dim = ws.dimensions
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        cell_n = 0
        sheet_formulas = 0
        sheet_errors: list[str] = []
        if max_row and max_col:
            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                for cell in row:
                    if cell.value is None:
                        continue
                    cell_n += 1
                    val = cell.value
                    if isinstance(val, str):
                        if FORMULA_RE.match(val):
                            sheet_formulas += 1
                            formula_count += 1
                        elif val.strip() in ERROR_VALUES:
                            error_cells += 1
                            sheet_errors.append(f"{cell.coordinate}={val.strip()}")
                            errors.append(
                                {
                                    "sheet": name,
                                    "cell": cell.coordinate,
                                    "message": "cached Excel error value in cell",
                                    "value": val.strip(),
                                }
                            )
        if cell_n > 0:
            non_empty += 1
        if ws.freeze_panes in (None, "A1"):
            warnings.append(
                {
                    "sheet": name,
                    "message": "no freeze panes; consider freezing header row",
                }
            )
        sheets_info.append(
            {
                "name": name,
                "dimensions": dim,
                "non_empty_cells": cell_n,
                "formulas": sheet_formulas,
                "error_samples": sheet_errors[:10],
            }
        )

    if non_empty == 0:
        errors.append({"message": "all sheets are empty"})

    wb.close()

    try:
        wb2 = load_workbook(path, data_only=True)
        for name in wb2.sheetnames:
            ws = wb2[name]
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            if not max_row or not max_col:
                continue
            for row in ws.iter_rows(
                min_row=1, max_row=min(max_row, 500), max_col=min(max_col, 50)
            ):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.strip() in ERROR_VALUES:
                        if not any(
                            e.get("cell") == cell.coordinate and e.get("sheet") == name
                            for e in errors
                        ):
                            errors.append(
                                {
                                    "sheet": name,
                                    "cell": cell.coordinate,
                                    "message": "data_only cache shows Excel error",
                                    "value": cell.value.strip(),
                                }
                            )
                            error_cells += 1
        wb2.close()
    except Exception:
        warnings.append({"message": "skipped data_only pass"})

    status = "ok" if not errors else "errors_found"
    return {
        "status": status,
        "file": str(path),
        "sheets": sheets_info,
        "formula_count": formula_count,
        "error_cell_count": error_cells,
        "error_count": len(errors),
        "errors": errors,
        "warnings": warnings,
    }


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Basic XLSX package / structure / error-value checks."
    )
    ap.add_argument("xlsx", type=Path)
    ap.add_argument("--json", action="store_true")
    args = ap.parse_args()
    if not args.xlsx.is_file():
        print(json.dumps({"status": "errors_found", "error": f"not found: {args.xlsx}"}))
        raise SystemExit(1)
    result = check_file(args.xlsx)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    raise SystemExit(0 if result["status"] == "ok" else 1)


if __name__ == "__main__":
    main()
